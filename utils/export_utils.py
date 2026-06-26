"""数据导出工具函数"""

import pandas as pd
import io
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ─── 列名映射 ────────────────────────────────────────

VENTILATION_EXPORT_MAP = {
    "segment_id": "管段编号",
    "airflow_m3h": "风量 Q（m³/h）",
    "width_mm": "宽度 a（mm）",
    "height_mm": "高度 b（mm）",
    "length_m": "长度 L（m）",
    "friction_pa_per_m": "单位长度摩擦阻力 R（Pa/m）",
    "zeta": "局部阻力系数 ζ",
    "风量 Q (m³/s)": "风量 q（m³/s）",
    "截面积 A (m²)": "截面积 A（m²）",
    "风速 v (m/s)": "风速 v（m/s）",
    "当量直径 De (m)": "当量直径 De（m）",
    "动压 Pd (Pa)": "动压 Pd（Pa）",
    "沿程阻力 Py (Pa)": "沿程阻力 Py（Pa）",
    "局部阻力 Pj (Pa)": "局部阻力 Pj（Pa）",
    "管段总阻力 P (Pa)": "管段总阻力 Pi（Pa）",
}

WATER_EXPORT_MAP = {
    "pipe_no": "管段编号",
    "load_kw": "负荷 QL（kW）",
    "flow_m3h": "水流量 G（m³/h）",
    "diameter_mm": "管道内径 D（mm）",
    "length_m": "长度 L（m）",
    "resistance_pa_per_m": "单位长度沿程阻力 R（Pa/m）",
    "local_zeta": "局部阻力系数 ζ",
    "流量 q (m³/s)": "流量 q（m³/s）",
    "截面积 A (m²)": "截面积 A（m²）",
    "流速 v (m/s)": "流速 v（m/s）",
    "动压 Pd (Pa)": "动压 Pd（Pa）",
    "沿程阻力 Py (Pa)": "沿程阻力 Py（Pa）",
    "局部阻力 Pj (Pa)": "局部阻力 Pj（Pa）",
    "管段总阻力 Pi (Pa)": "管段总阻力 Pi（Pa）",
    "流速校核": "流速校核",
}


# ─── 通用函数 ────────────────────────────────────────

def get_csv_bytes(df):
    """返回 CSV 文件的 bytes"""
    buf = io.BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    return buf.getvalue()


def rename_export_columns(df, col_map):
    """按映射重命名并选取列，仅保留映射中存在的列"""
    available = {k: v for k, v in col_map.items() if k in df.columns}
    return df[list(available.keys())].rename(columns=available)


def _apply_excel_format(ws, df):
    """对 worksheet 设置表头加粗和自动列宽"""
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        col_width = max(
            len(str(col_name)),
            df[col_name].astype(str).map(len).max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(col_width + 4, 28)


def export_formatted_excel(data_df, summary_rows, sheet_name="管段计算结果",
                           summary_sheet_name="系统汇总"):
    """
    生成带格式的 Excel 文件 bytes。

    Parameters
    ----------
    data_df : DataFrame
        管段计算结果（列名已为中文）
    summary_rows : list[tuple[str, str]]
        汇总数据，每项为 (标签, 数值)，例如 [("系统总阻力", "60.70 Pa")]
    sheet_name : str
        数据 sheet 名
    summary_sheet_name : str
        汇总 sheet 名
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # 管段计算结果 sheet
        data_df.to_excel(writer, sheet_name=sheet_name, index=False)
        _apply_excel_format(writer.sheets[sheet_name], data_df)

        # 系统汇总 sheet
        summ_df = pd.DataFrame(summary_rows, columns=["项目", "数值"])
        summ_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)
        _apply_excel_format(writer.sheets[summary_sheet_name], summ_df)

    return buf.getvalue()
